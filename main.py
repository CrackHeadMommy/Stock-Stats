import json #for cleaner code stock dictionary is saved in a separete file
import requests #needed to access html code of finance.yahoo.com
from bs4 import BeautifulSoup #web scarping tool for data extraction from html code
import time #used to delay scraping aas to not get blocked by yahoo, also for timestamping each scraped stock value
from openpyxl import Workbook, load_workbook #allows to open/create/modify excel files where I record scraped values

# Load stock name/ticker dictionary from JSON file
with open("stocks.json", "r") as f:
    stock_dictionary = json.load(f)

# Define a custom class to store stock tickers
class ListTicker:
    def __init__(self):
        #empty listr
        self.tickers = []

    def add_ticker(self, ticker):
        #adds new tickers to list
        if ticker.upper() not in self.tickers:
            self.tickers.append(ticker.upper())

#user prompt
user_input = input("Enter up to 3 stocks (names or tickers, comma-separated): ")
inputs = [x.strip() for x in user_input.split(",") if x.strip()]

#user inputed list of tickers
input_tickers = ListTicker()

for entry in inputs:
    #checks if a valid ticker
    if entry.upper() in stock_dictionary.values():
        input_tickers.add_ticker(entry.upper())
    #checks if needs conversion
    elif entry in stock_dictionary:
        input_tickers.add_ticker(stock_dictionary[entry])
    else:
        #asks if the user wants to proceed with input as a unknown ticker
        check = input(f"'{entry}' not found. Use as ticker anyway? (y/n): ").lower()
        if check == 'y':
            input_tickers.add_ticker(entry.upper())

# Retrieve the final list of tickers
tickers = input_tickers.tickers

#no tickers
if not tickers:
    print("No valid tickers given. Exiting.")
    exit()

#tries to load excel file
try:
    wb = load_workbook("stock_prices.xlsx")
    ws = wb.active
except FileNotFoundError:
#creates excel file if none found
    wb = Workbook()
    ws = wb.active
    ws.append(["Time"] + tickers)

def get_price(ticker):
    #gets stock price for each ticker from url
    url = f"https://finance.yahoo.com/quote/{ticker}"
    headers = {"User-Agent": "Mozilla/5.0"} #bot masking     Generated using ChatGPT : prompt: "How can i make my attempts in web scraping look like a normal browser request in pyhton code?"

    #looks through html code and finds price element
    resp = requests.get(url, headers=headers)
    soup = BeautifulSoup(resp.text, "html.parser")
    price = soup.find("span", {"data-testid": "qsp-price"})
    return price.text.strip()


def scraped_data():
    #constantly scrapes price value
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")#records time
    prices = [get_price(t) for t in tickers]
    print(f"[{timestamp}] " + ", ".join(f"{t}: {p}" for t, p in zip(tickers, prices)))#displays in the terminal
    ws.append([timestamp] + prices)
    wb.save("stock_prices.xlsx")

while True:
    #scraping interval
    scraped_data()
    time.sleep(15)
